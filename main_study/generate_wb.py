from openpyxl import load_workbook, Workbook
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import os 
import shutil 

tweet_df = pd.read_csv("test_tweets.csv", dtype={'coordinates': str})

src = r"survey_template.xlsx"
num_participants = 7
num_tweets = 1 # how many tweets per participant

for i in range(num_participants):
    for i in range(num_tweets):
        # pick a tweet (randomly + make sure each tweet only gets picked twice)
        
        # get screen shot of tweet
        
        # get images in tweet if any
        
        pass  
        
        # insert images into new file
        
    dest = r"survey_copies/copy_" + str(i) + ".xlsx"
    path = shutil.copyfile(src, dest)
    
